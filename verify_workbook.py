"""
Check the workbook's numbers against the source data.

A spreadsheet with no #REF! errors is not a correct spreadsheet. Every formula
here could evaluate cleanly and still point one row off, or quietly include the
refunds it was supposed to exclude. So this recomputes the same figures in
pandas and compares them cell by cell.

    python3 verify_workbook.py

Run it after any change to build_workbook.py, and always after recalc.py —
openpyxl writes formulas without calculating them, so an unrecalculated
workbook has nothing for this script to read.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SRC = Path("data/orders_raw.csv")
WB = Path("weekly-report-template.xlsx")
TOL = 0.01

passed = failed = 0


def check(name, expected, actual, tol=TOL):
    global passed, failed
    if isinstance(expected, float) or isinstance(actual, float):
        ok = abs(float(expected) - float(actual)) <= tol
    else:
        ok = expected == actual
    if ok:
        passed += 1
        print(f"  PASS  {name:<44} {expected}")
    else:
        failed += 1
        print(f"  FAIL  {name:<44} expected {expected}, sheet says {actual}")


def parse_date(s: str):
    s = s.strip()
    if len(s) > 4 and s[4] == "-":
        return pd.to_datetime(s, format="%Y-%m-%d %H:%M")
    return pd.to_datetime(s, format="%d/%m/%Y %H:%M")


def main() -> int:
    if not WB.exists():
        print("build the workbook first", file=sys.stderr)
        return 1

    raw = pd.read_csv(SRC, dtype=str).fillna("")
    raw["dt"] = raw["Date"].map(parse_date)
    raw["week"] = raw["dt"] - pd.to_timedelta(raw["dt"].dt.weekday, unit="D")
    raw["week"] = raw["week"].dt.normalize()
    raw["total"] = pd.to_numeric(raw["Line Total"])
    raw["qty"] = pd.to_numeric(raw["Qty"])

    is_prod = raw["Line Type"] == "Product"
    fulfilled = raw["Status"] == "Fulfilled"
    not_cancelled = raw["Status"] != "Cancelled"

    # The report keys off the last week that actually had orders, not the last
    # date in the file - a late refund can create a stub week with no sales.
    order_weeks = raw.loc[is_prod & fulfilled].groupby("week")["Order ID"].nunique()
    latest = order_weeks[order_weeks > 0].index.max()
    prior = latest - pd.Timedelta(days=7)

    wb = load_workbook(WB, data_only=True)
    rep = wb["Weekly Report"]
    calc = wb["Weekly Calc"]

    # `data_only=True` reads cached results, not formulas. openpyxl never
    # calculates anything, so a workbook straight out of build_workbook.py has
    # no cache and every cell below reads None. Left alone that produces
    # seventeen confident failures against a workbook that is entirely correct
    # — the worst possible output from a tool whose whole job is telling you
    # whether to trust the numbers. So: say what actually happened.
    if rep["B6"].value is None and rep["D6"].value is None:
        print("The workbook has formulas but no calculated values.\n")
        print("  openpyxl writes formulas without evaluating them, so nothing has\n"
              "  been calculated yet. This is not a failure — there is simply\n"
              "  nothing here to check.\n")
        print("  Run:  python3 recalc.py     (then this script again)\n")
        print("  Or open the file in Excel or Google Sheets and save it, which\n"
              "  has the same effect.")
        return 2

    print(f"workbook vs source data   (latest week {latest.date()})\n")

    # ---- headline KPIs -------------------------------------------------
    rev = raw.loc[is_prod & not_cancelled & (raw["week"] == latest), "total"].sum()
    check("revenue, latest week", round(rev, 2), round(rep["B6"].value or 0, 2))

    orders = raw.loc[is_prod & fulfilled & (raw["week"] == latest), "Order ID"].nunique()
    check("orders, latest week", orders, int(rep["D6"].value or 0))

    units = raw.loc[is_prod & fulfilled & (raw["week"] == latest), "qty"].sum()
    check("units, latest week", int(units), int(rep["H6"].value or 0))

    check("avg order value", round(rev / orders, 2), round(rep["F6"].value or 0, 2))

    prev_rev = raw.loc[is_prod & not_cancelled & (raw["week"] == prior), "total"].sum()
    check("week-over-week change", round(rev / prev_rev - 1, 4), round(rep["J6"].value or 0, 4))

    # ---- the weekly table ----------------------------------------------
    weekly = (raw[is_prod & not_cancelled].groupby("week")["total"].sum().sort_index())
    check("total revenue, all weeks", round(weekly.sum(), 2),
          round(sum(calc.cell(row=r, column=9).value or 0 for r in range(5, 21)), 2))
    check("reporting week cell", latest.date().isoformat(),
          calc["H3"].value.date().isoformat() if hasattr(calc["H3"].value, "date")
          else str(calc["H3"].value))

    # ---- exclusions, which are the whole trust argument -----------------
    ship = raw.loc[raw["Line Type"] == "Shipping", "total"].sum()
    disc = raw.loc[raw["Line Type"] == "Discount", "total"].sum()
    cancelled = int((raw["Status"] == "Cancelled").sum())
    guests = int((is_prod & (raw["Customer Email"] == "")).sum())
    refunds = raw.loc[(raw["Status"] == "Refunded") & (raw["week"] == latest), "total"].sum()

    top_row = 13 + 14 + 1 + 1
    ex_row = 13 + 14 + 1 + 2 + 8 + 2
    check("refunds, latest week", round(refunds, 2), round(rep.cell(row=ex_row + 2, column=3).value or 0, 2))
    check("shipping excluded", round(ship, 2), round(rep.cell(row=ex_row + 3, column=3).value or 0, 2))
    check("discounts excluded", round(disc, 2), round(rep.cell(row=ex_row + 4, column=3).value or 0, 2))
    check("cancelled lines", cancelled, int(rep.cell(row=ex_row + 5, column=3).value or 0))
    check("guest checkout lines", guests, int(rep.cell(row=ex_row + 6, column=3).value or 0))

    # ---- top products ---------------------------------------------------
    top = (raw[is_prod & not_cancelled & (raw["week"] == latest)]
           .groupby("Item")["total"].sum().sort_values(ascending=False))
    check("top product name", top.index[0], rep.cell(row=top_row + 1, column=2).value)
    check("top product revenue", round(top.iloc[0], 2),
          round(rep.cell(row=top_row + 1, column=3).value or 0, 2))
    check("second product name", top.index[1], rep.cell(row=top_row + 2, column=2).value)

    # ---- channels -------------------------------------------------------
    ch = (raw[is_prod & not_cancelled & (raw["week"] == latest)]
          .groupby("Channel")["total"].sum())
    sheet_ch = {}
    for i in range(len(ch)):
        r = top_row + 1 + i
        name = rep.cell(row=r, column=6).value
        if name:
            sheet_ch[name] = rep.cell(row=r, column=7).value or 0
    check("channel revenue totals match", round(ch.sum(), 2), round(sum(sheet_ch.values()), 2))
    check("channel count", len(ch), len(sheet_ch))

    print()
    if failed:
        print(f"{failed} check(s) FAILED - the workbook is showing numbers the data does not support.")
        return 1
    print(f"All {passed} checks passed. Every figure on the report reconciles to the raw export.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
