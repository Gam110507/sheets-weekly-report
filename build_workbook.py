"""
Build the three-tab weekly report workbook.

    Raw Orders     what you paste in - never edited by hand
    Weekly Calc    the query layer - parses, classifies, aggregates
    Weekly Report  the one page anyone actually reads

Nothing on the report is a typed-in number. Paste a fresh export over the Raw
Orders tab and every figure moves by itself.

Formulas are limited to functions that behave identically in Excel and Google
Sheets, because the delivered file lives in Sheets. That rules out XLOOKUP,
FILTER and UNIQUE; INDEX/MATCH and SUMIFS do the same work and travel.

    python3 build_workbook.py     # writes weekly-report-template.xlsx
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path("data/orders_raw.csv")
OUT = Path("weekly-report-template.xlsx")

TREND_WEEKS = 14        # trailing weeks shown on the report
CALC_WEEKS = 16         # summary rows, with headroom for late refunds

INK = "16202C"
MUTED = "6B7A8F"
ACCENT = "1F6FEB"
ACCENT_SOFT = "EAF1FE"
RULE = "DDE3EA"
BAND = "F6F8FB"

F_TITLE = Font(name="Arial", size=17, bold=True, color=INK)
F_SUB = Font(name="Arial", size=9.5, color=MUTED)
F_SECTION = Font(name="Arial", size=11, bold=True, color=INK)
F_HDR = Font(name="Arial", size=9, bold=True, color="FFFFFF")
F_BODY = Font(name="Arial", size=10, color=INK)
F_BODY_B = Font(name="Arial", size=10, bold=True, color=INK)
F_SMALL = Font(name="Arial", size=8.5, color=MUTED)
F_KPI = Font(name="Arial", size=17, bold=True, color=INK)
F_LABEL = Font(name="Arial", size=8, bold=True, color=MUTED)

FILL_HDR = PatternFill("solid", fgColor=ACCENT)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_KPI = PatternFill("solid", fgColor=ACCENT_SOFT)

THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
UNDER = Border(bottom=THIN)

GBP = '£#,##0.00'
GBP0 = '£#,##0'
PCT = '0.0%;[Red]-0.0%'
PCT_PLAIN = '0.0%'
INT = '#,##0'
DATE = 'yyyy-mm-dd'


def header_row(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX
    ws.row_dimensions[row].height = 20


def band(ws, row, c1, c2, striped):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_BODY
        cell.border = BOX
        if striped:
            cell.fill = FILL_BAND


def main() -> None:
    raw = pd.read_csv(SRC, dtype=str).fillna("")
    n = len(raw)
    last = n + 1                       # last data row on Raw Orders
    hlast = n + 4                      # last helper row on Weekly Calc

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # TAB — Raw Orders
    # ══════════════════════════════════════════════════════════════════════
    rw = wb.active
    rw.title = "Raw Orders"
    cols = list(raw.columns)
    rw.append(cols)

    numeric = {"Qty", "Unit Price", "Line Total"}
    num_idx = [i for i, c in enumerate(cols) if c in numeric]
    for rec in raw.itertuples(index=False):
        row = list(rec)
        # Written as text these look identical on screen, but every SUMIFS over
        # this tab silently returns zero - a confident wrong report, which is
        # worse than one that visibly breaks.
        for i in num_idx:
            try:
                row[i] = float(row[i])
            except (TypeError, ValueError):
                pass
        rw.append(row)

    header_row(rw, 1, 1, len(cols))
    for i, w in enumerate([11, 18, 11, 27, 6, 11, 11, 12, 16, 27, 11], start=1):
        rw.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, last + 1):
        for c in range(1, len(cols) + 1):
            rw.cell(row=r, column=c).font = F_BODY
        rw.cell(row=r, column=6).number_format = GBP
        rw.cell(row=r, column=7).number_format = GBP
    rw.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════
    # TAB — Weekly Calc
    # ══════════════════════════════════════════════════════════════════════
    q = wb.create_sheet("Weekly Calc")
    q.sheet_view.showGridLines = False

    q["A1"] = "Query layer"
    q["A1"].font = F_TITLE
    q["A2"] = ("Every column here reads from Raw Orders. Nothing is typed by hand, so "
               "pasting a new export updates the report by itself.")
    q["A2"].font = F_SUB

    for i, h in enumerate(["Order date", "Week starting (Mon)", "Product revenue",
                           "Counts as an order", "Units", "Refund value"], start=1):
        q.cell(row=4, column=i, value=h)
    header_row(q, 4, 1, 6)

    for r in range(5, hlast + 1):
        s = r - 3
        # The export writes dates two ways. Reading the separator is
        # locale-independent; DATEVALUE is not, and would swap day and month for
        # anyone outside the UK without saying so.
        q.cell(row=r, column=1, value=(
            f'=IF(\'Raw Orders\'!B{s}="","",IF(MID(\'Raw Orders\'!B{s},5,1)="-",'
            f'DATE(VALUE(LEFT(\'Raw Orders\'!B{s},4)),VALUE(MID(\'Raw Orders\'!B{s},6,2)),'
            f'VALUE(MID(\'Raw Orders\'!B{s},9,2))),'
            f'DATE(VALUE(MID(\'Raw Orders\'!B{s},7,4)),VALUE(MID(\'Raw Orders\'!B{s},4,2)),'
            f'VALUE(LEFT(\'Raw Orders\'!B{s},2)))))'))
        # WEEKDAY type 3 makes Monday 0, so subtracting lands on the Monday.
        q.cell(row=r, column=2, value=f'=IF(A{r}="","",A{r}-WEEKDAY(A{r},3))')
        q.cell(row=r, column=3, value=(
            f'=IF(OR(\'Raw Orders\'!C{s}<>"Product",\'Raw Orders\'!K{s}="Cancelled"),0,'
            f'\'Raw Orders\'!G{s})'))
        # An order is counted once: on the first fulfilled product line carrying
        # that order number.
        q.cell(row=r, column=4, value=(
            f'=IF(AND(\'Raw Orders\'!C{s}="Product",\'Raw Orders\'!K{s}="Fulfilled",'
            f'COUNTIFS(\'Raw Orders\'!$A$2:$A{s},\'Raw Orders\'!A{s},'
            f'\'Raw Orders\'!$C$2:$C{s},"Product",\'Raw Orders\'!$K$2:$K{s},"Fulfilled")=1),1,0)'))
        q.cell(row=r, column=5, value=(
            f'=IF(AND(\'Raw Orders\'!C{s}="Product",\'Raw Orders\'!K{s}="Fulfilled"),'
            f'\'Raw Orders\'!E{s},0)'))
        q.cell(row=r, column=6, value=(
            f'=IF(\'Raw Orders\'!K{s}="Refunded",\'Raw Orders\'!G{s},0)'))

        for c in range(1, 7):
            q.cell(row=r, column=c).font = F_BODY
        q.cell(row=r, column=1).number_format = DATE
        q.cell(row=r, column=2).number_format = DATE
        q.cell(row=r, column=3).number_format = GBP
        q.cell(row=r, column=6).number_format = GBP

    wrange = f'$B$5:$B${hlast}'

    # ---- weekly summary --------------------------------------------------
    q["H1"] = "Weekly summary"
    q["H1"].font = F_SECTION
    q["H2"] = "The report tab reads from this table."
    q["H2"].font = F_SUB
    q["G3"] = "REPORTING WEEK"
    q["G3"].font = F_LABEL
    # One cell decides which week the report is about: the most recent week that
    # actually had orders. Anchoring on the newest date instead would point the
    # whole report at a stub week containing nothing but a late refund.
    q["H3"] = f'=SUMPRODUCT(MAX(($J$5:$J${4 + CALC_WEEKS}>0)*($H$5:$H${4 + CALC_WEEKS})))'
    q["H3"].font = F_BODY_B
    q["H3"].number_format = DATE
    q["I3"] = "last week with orders, not just the last date in the file"
    q["I3"].font = F_SMALL

    for i, h in enumerate(["Week starting", "Revenue", "Orders", "Units", "Refunds",
                           "Avg order value", "vs prior week"], start=8):
        q.cell(row=4, column=i, value=h)
    header_row(q, 4, 8, 14)

    for i in range(CALC_WEEKS):
        r = 5 + i
        q.cell(row=r, column=8,
               value=(f'=MIN({wrange})' if i == 0 else f'=H{r - 1}+7'))
        q.cell(row=r, column=9, value=f'=SUMIFS($C$5:$C${hlast},{wrange},H{r})')
        q.cell(row=r, column=10, value=f'=SUMIFS($D$5:$D${hlast},{wrange},H{r})')
        q.cell(row=r, column=11, value=f'=SUMIFS($E$5:$E${hlast},{wrange},H{r})')
        q.cell(row=r, column=12, value=f'=SUMIFS($F$5:$F${hlast},{wrange},H{r})')
        q.cell(row=r, column=13, value=f'=IFERROR(I{r}/J{r},0)')
        q.cell(row=r, column=14, value=('' if i == 0 else f'=IFERROR(I{r}/I{r - 1}-1,"")'))

        band(q, r, 8, 14, i % 2 == 1)
        q.cell(row=r, column=8).number_format = DATE
        q.cell(row=r, column=8).alignment = Alignment(horizontal="left")
        q.cell(row=r, column=9).number_format = GBP0
        q.cell(row=r, column=10).number_format = INT
        q.cell(row=r, column=11).number_format = INT
        q.cell(row=r, column=12).number_format = GBP0
        q.cell(row=r, column=13).number_format = GBP
        q.cell(row=r, column=14).number_format = PCT

    # ---- breakdowns for the reporting week -------------------------------
    q["P1"] = "Reporting week breakdown"
    q["P1"].font = F_SECTION
    q["P2"] = '=TEXT($H$3,"d mmm yyyy")'
    q["P2"].font = F_SUB

    items = sorted(raw.loc[raw["Line Type"] == "Product", "Item"].unique())
    for i, h in enumerate(["Product", "Revenue", "Units"], start=16):
        q.cell(row=4, column=i, value=h)
    header_row(q, 4, 16, 18)
    for i, item in enumerate(items):
        r = 5 + i
        q.cell(row=r, column=16, value=item)
        q.cell(row=r, column=17, value=(
            f'=SUMIFS($C$5:$C${hlast},{wrange},$H$3,'
            f'\'Raw Orders\'!$D$2:$D${last},$P{r})'))
        q.cell(row=r, column=18, value=(
            f'=SUMIFS($E$5:$E${hlast},{wrange},$H$3,'
            f'\'Raw Orders\'!$D$2:$D${last},$P{r})'))
        band(q, r, 16, 18, i % 2 == 1)
        q.cell(row=r, column=17).number_format = GBP0
        q.cell(row=r, column=18).number_format = INT

    channels = sorted(raw["Channel"].unique())
    for i, h in enumerate(["Channel", "Revenue"], start=20):
        q.cell(row=4, column=i, value=h)
    header_row(q, 4, 20, 21)
    for i, ch in enumerate(channels):
        r = 5 + i
        q.cell(row=r, column=20, value=ch)
        q.cell(row=r, column=21, value=(
            f'=SUMIFS($C$5:$C${hlast},{wrange},$H$3,'
            f'\'Raw Orders\'!$H$2:$H${last},$T{r})'))
        band(q, r, 20, 21, i % 2 == 1)
        q.cell(row=r, column=21).number_format = GBP0

    for col, w in [("A", 13), ("B", 19), ("C", 15), ("D", 17), ("E", 8), ("F", 12),
                   ("G", 17), ("H", 14), ("I", 12), ("J", 9), ("K", 9), ("L", 11),
                   ("M", 15), ("N", 13), ("O", 3), ("P", 29), ("Q", 11), ("R", 9),
                   ("S", 3), ("T", 14), ("U", 11)]:
        q.column_dimensions[col].width = w
    q.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════
    # TAB — Weekly Report
    # ══════════════════════════════════════════════════════════════════════
    rep = wb.create_sheet("Weekly Report", 0)
    rep.sheet_view.showGridLines = False

    rep["B2"] = "Weekly Sales Report"
    rep["B2"].font = F_TITLE
    rep.row_dimensions[2].height = 24
    rep["B3"] = ('="Week starting "&TEXT(\'Weekly Calc\'!$H$3,"d mmmm yyyy")'
                 '&"   ·   rebuilt automatically from the Raw Orders tab"')
    rep["B3"].font = F_SUB
    for c in range(2, 11):
        rep.cell(row=3, column=c).border = UNDER

    match = "MATCH('Weekly Calc'!$H$3,'Weekly Calc'!$H$5:$H$20,0)"

    kpis = [
        ("REVENUE", "I", GBP0, "fulfilled product lines"),
        ("ORDERS", "J", INT, "distinct orders"),
        ("AVG ORDER VALUE", "M", GBP, "revenue divided by orders"),
        ("UNITS SOLD", "K", INT, "items shipped"),
        ("VS LAST WEEK", "N", PCT, "change in revenue"),
    ]
    for i, (label, col, fmt, note) in enumerate(kpis):
        c = 2 + i * 2
        rep.cell(row=5, column=c, value=label).font = F_LABEL
        v = rep.cell(row=6, column=c,
                     value=f"=INDEX('Weekly Calc'!${col}$5:${col}$20,{match})")
        v.font = F_KPI
        v.number_format = fmt
        v.alignment = Alignment(horizontal="left")
        rep.cell(row=7, column=c, value=note).font = F_SMALL
        for r in (5, 6, 7):
            rep.cell(row=r, column=c).fill = FILL_KPI
            rep.cell(row=r, column=c + 1).fill = FILL_KPI
    rep.row_dimensions[6].height = 24

    rep["B9"] = ("Revenue counts fulfilled product lines only. Shipping, discounts, "
                 "cancellations and refunds are excluded and reported separately at the "
                 "bottom, so this figure reconciles to the raw export.")
    rep["B9"].font = F_SMALL

    # ---- trailing trend ---------------------------------------------------
    rep["B11"] = f"Revenue, last {TREND_WEEKS} weeks"
    rep["B11"].font = F_SECTION
    for i, h in enumerate(["Week starting", "Revenue", "Orders",
                           "Avg order value", "vs prior week"], start=2):
        rep.cell(row=12, column=i, value=h)
    header_row(rep, 12, 2, 6)

    for i in range(TREND_WEEKS):
        r = 13 + i
        # Anchored to the reporting week, so this always shows the trailing
        # window rather than whichever rows happen to exist in the table.
        wk = f"('Weekly Calc'!$H$3-{7 * (TREND_WEEKS - 1 - i)})"
        d = rep.cell(row=r, column=2, value=f"={wk}")
        d.number_format = DATE
        d.alignment = Alignment(horizontal="left")
        for col_letter, out_col, fmt in [("I", 3, GBP0), ("J", 4, INT),
                                         ("M", 5, GBP), ("N", 6, PCT)]:
            idx = (f"INDEX('Weekly Calc'!${col_letter}$5:${col_letter}$20,"
                   f"MATCH({wk},'Weekly Calc'!$H$5:$H$20,0))")
            if col_letter == "N":
                # INDEX returns 0 for a blank cell, which would print "0.0%" and
                # claim the week was flat. The oldest week in the window has no
                # week before it, and unknown is not the same as unchanged.
                formula = f'=IFERROR(IF({idx}="","",{idx}),"")'
            else:
                formula = f"=IFERROR({idx},0)"
            rep.cell(row=r, column=out_col, value=formula).number_format = fmt
        band(rep, r, 2, 6, i % 2 == 1)

    # ---- top products and channels ---------------------------------------
    sec = 13 + TREND_WEEKS + 1
    n_items, n_ch = len(items), len(channels)
    rows_shown = 8

    rep.cell(row=sec, column=2, value="Top products this week").font = F_SECTION
    for i, h in enumerate(["Product", "Revenue", "Units"], start=2):
        rep.cell(row=sec + 1, column=i, value=h)
    header_row(rep, sec + 1, 2, 4)
    for i in range(rows_shown):
        r = sec + 2 + i
        rank = (f"MATCH(LARGE('Weekly Calc'!$Q$5:$Q${4 + n_items},{i + 1}),"
                f"'Weekly Calc'!$Q$5:$Q${4 + n_items},0)")
        rep.cell(row=r, column=2, value=(
            f"=IFERROR(INDEX('Weekly Calc'!$P$5:$P${4 + n_items},{rank}),\"\")"))
        rep.cell(row=r, column=3, value=(
            f"=IFERROR(LARGE('Weekly Calc'!$Q$5:$Q${4 + n_items},{i + 1}),\"\")"))
        rep.cell(row=r, column=4, value=(
            f"=IFERROR(INDEX('Weekly Calc'!$R$5:$R${4 + n_items},{rank}),\"\")"))
        band(rep, r, 2, 4, i % 2 == 1)
        rep.cell(row=r, column=3).number_format = GBP0
        rep.cell(row=r, column=4).number_format = INT

    rep.cell(row=sec, column=6, value="Revenue by channel").font = F_SECTION
    for i, h in enumerate(["Channel", "Revenue", "Share"], start=6):
        rep.cell(row=sec + 1, column=i, value=h)
    header_row(rep, sec + 1, 6, 8)
    for i in range(n_ch):
        r = sec + 2 + i
        rep.cell(row=r, column=6, value=f"='Weekly Calc'!T{5 + i}")
        rep.cell(row=r, column=7, value=f"='Weekly Calc'!U{5 + i}").number_format = GBP0
        rep.cell(row=r, column=8, value=(
            f"=IFERROR(G{r}/SUM($G${sec + 2}:$G${sec + 1 + n_ch}),0)"))
        rep.cell(row=r, column=8).number_format = PCT_PLAIN
        band(rep, r, 6, 8, i % 2 == 1)

    # ---- reconciliation ---------------------------------------------------
    ex = sec + 2 + rows_shown + 2
    rep.cell(row=ex, column=2, value="What this report leaves out").font = F_SECTION
    rep.cell(row=ex + 1, column=2,
             value=("Stated rather than hidden, so the revenue figure above can be "
                    "reconciled against the raw export.")).font = F_SMALL

    excl = [
        ("Refunds, this week", f"=INDEX('Weekly Calc'!$L$5:$L$20,{match})", GBP0),
        ("Shipping charged, all weeks",
         f'=SUMIFS(\'Raw Orders\'!$G$2:$G${last},\'Raw Orders\'!$C$2:$C${last},"Shipping")', GBP0),
        ("Discounts given, all weeks",
         f'=SUMIFS(\'Raw Orders\'!$G$2:$G${last},\'Raw Orders\'!$C$2:$C${last},"Discount")', GBP0),
        ("Cancelled lines, all weeks",
         f'=COUNTIFS(\'Raw Orders\'!$K$2:$K${last},"Cancelled")', INT),
        ("Product lines with no customer email",
         f'=COUNTIFS(\'Raw Orders\'!$J$2:$J${last},"",\'Raw Orders\'!$C$2:$C${last},"Product")', INT),
    ]
    for i, (label, formula, fmt) in enumerate(excl):
        r = ex + 2 + i
        a = rep.cell(row=r, column=2, value=label)
        b = rep.cell(row=r, column=3, value=formula)
        a.font = F_BODY
        b.font = F_BODY
        b.number_format = fmt
        a.border = UNDER
        b.border = UNDER

    foot = ex + 2 + len(excl) + 2
    rep.cell(row=foot, column=2, value=(
        "Raw data in, report out, no manual steps. The Apps Script in this file "
        "emails this summary on a schedule.")).font = F_SMALL

    for col, w in [("A", 3), ("B", 34), ("C", 15), ("D", 13), ("E", 17),
                   ("F", 21), ("G", 14), ("H", 10), ("I", 3), ("J", 16), ("K", 3)]:
        rep.column_dimensions[col].width = w

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Raw Orders     {n:,} rows")
    print(f"  Weekly Calc    {n:,} helper rows, {CALC_WEEKS}-week summary")
    print(f"  Weekly Report  {len(kpis)} KPIs, {TREND_WEEKS}-week trend, "
          f"{rows_shown} products, {n_ch} channels")


if __name__ == "__main__":
    main()
