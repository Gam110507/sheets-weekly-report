"""
Render the scheduled email exactly as the Apps Script would send it.

The HTML template here is a direct port of renderEmail_() in
apps-script/Code.gs, and the numbers are read out of the built workbook rather
than typed in. So the preview is the real thing with real figures, not a mockup.

Two deliberate differences, both because this runs outside Google: the "open
the full report" link has no Sheet URL to point at, and the plain-text
alternative body that Code.gs also sends is not rendered. Everything visible in
the HTML is what arrives.

    python3 render_email_preview.py     # writes screenshots/email-preview.html
"""

from pathlib import Path

from openpyxl import load_workbook

WB = Path("weekly-report-template.xlsx")
OUT = Path("screenshots/email-preview.html")


def money(v):
    return "£" + f"{round(float(v or 0)):,}"


def pct(v):
    n = (float(v or 0)) * 100
    return f"{'+' if n >= 0 else ''}{n:.1f}%"


def main() -> None:
    wb = load_workbook(WB, data_only=True)
    rep = wb["Weekly Report"]
    calc = wb["Weekly Calc"]

    week = calc["H3"].value
    week_label = week.strftime("%-d %b %Y") if hasattr(week, "strftime") else str(week)

    k = {
        "revenue": rep["B6"].value,
        "orders": rep["D6"].value,
        "aov": rep["F6"].value,
        "wow": rep["J6"].value,
    }

    sec = 13 + 14 + 1                       # section header row
    products, channels = [], []
    for i in range(8):
        r = sec + 2 + i
        name = rep.cell(row=r, column=2).value
        if name:
            products.append((name, rep.cell(row=r, column=3).value,
                             rep.cell(row=r, column=4).value))
    for i in range(5):          # Code.gs reads 5 channel rows; match it exactly
        r = sec + 2 + i
        name = rep.cell(row=r, column=6).value
        if name:
            channels.append((name, rep.cell(row=r, column=7).value,
                             rep.cell(row=r, column=8).value))

    up = float(k["wow"] or 0) >= 0
    trend = "#157F51" if up else "#B42318"

    def kpi(label, value, note):
        return (
            '<td style="padding:14px 18px;background:#EAF1FE;border-radius:8px;vertical-align:top;">'
            f'<div style="font:600 10px Arial;letter-spacing:.06em;color:#6B7A8F;text-transform:uppercase;">{label}</div>'
            f'<div style="font:700 22px Arial;color:#16202C;padding-top:4px;">{value}</div>'
            f'<div style="font:400 11px Arial;color:#6B7A8F;padding-top:2px;">{note}</div>'
            '</td><td style="width:10px;"></td>')

    def th(labels):
        cells = "".join(
            '<th style="padding:7px 10px;font:700 10px Arial;letter-spacing:.05em;'
            'text-transform:uppercase;color:#FFFFFF;background:#1F6FEB;text-align:'
            f'{"left" if i == 0 else "right"};">{l}</th>'
            for i, l in enumerate(labels))
        return f"<tr>{cells}</tr>"

    def rows(data):
        out = []
        for i, (name, rev, third) in enumerate(data):
            bg = "#F6F8FB" if i % 2 else "#FFFFFF"
            t = (f"{float(third) * 100:.1f}%" if isinstance(third, float) and third <= 1
                 else f"{round(float(third or 0))}")
            out.append(
                f'<tr style="background:{bg};">'
                f'<td style="padding:7px 10px;font:400 13px Arial;color:#16202C;">{name}</td>'
                f'<td style="padding:7px 10px;font:400 13px Arial;color:#16202C;text-align:right;">{money(rev)}</td>'
                f'<td style="padding:7px 10px;font:400 13px Arial;color:#6B7A8F;text-align:right;">{t}</td>'
                "</tr>")
        return "".join(out)

    html = (
        '<div style="background:#F6F8FB;padding:24px;font-family:Arial,sans-serif;">'
        '<div style="max-width:640px;margin:0 auto;background:#FFFFFF;border:1px solid #DDE3EA;'
        'border-radius:12px;padding:26px 28px;">'
        '<div style="font:700 20px Arial;color:#16202C;">Weekly Sales Report</div>'
        f'<div style="font:400 12px Arial;color:#6B7A8F;padding:4px 0 18px;border-bottom:1px solid #DDE3EA;">Week starting {week_label}</div>'
        '<table cellpadding="0" cellspacing="0" style="margin:18px 0 6px;"><tr>'
        + kpi("Revenue", money(k["revenue"]),
              f'<span style="color:{trend};font-weight:700;">{pct(k["wow"])}</span> vs last week')
        + kpi("Orders", f'{round(float(k["orders"] or 0))}', "distinct orders")
        + kpi("Avg order", money(k["aov"]), "per order")
        + "</tr></table>"
        '<div style="font:700 13px Arial;color:#16202C;padding:22px 0 8px;">Top products</div>'
        '<table cellpadding="0" cellspacing="0" width="100%" style="border-collapse:collapse;">'
        + th(["Product", "Revenue", "Units"]) + rows(products) + "</table>"
        '<div style="font:700 13px Arial;color:#16202C;padding:22px 0 8px;">By channel</div>'
        '<table cellpadding="0" cellspacing="0" width="100%" style="border-collapse:collapse;">'
        + th(["Channel", "Revenue", "Share"]) + rows(channels) + "</table>"
        '<div style="font:400 11px Arial;color:#6B7A8F;padding:20px 0 0;border-top:1px solid #DDE3EA;margin-top:22px;">'
        "Revenue counts fulfilled product lines only. Shipping, discounts, cancellations "
        "and refunds are excluded and listed separately in the Sheet, so this figure "
        "reconciles to the raw export.<br><br>"
        '<a href="#" style="color:#1F6FEB;text-decoration:none;font-weight:700;">Open the full report &rarr;</a>'
        "</div></div></div>")

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(
        "<!doctype html><meta charset='utf-8'>"
        "<body style='margin:0;background:#F6F8FB;'>" + html + "</body>",
        encoding="utf-8")
    print(f"wrote {OUT}   week {week_label}, revenue {money(k['revenue'])}, "
          f"{len(products)} products, {len(channels)} channels")


if __name__ == "__main__":
    main()
