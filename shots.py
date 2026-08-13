"""
Render the screenshots used in README.md.

    python3 shots.py

Three images, all generated from the current files rather than cropped by hand:

    screenshots/weekly-report.png     the report tab as the client sees it
    screenshots/raw-orders-before.png the top of the messy export
    screenshots/email-preview.png     the scheduled email

Why bother scripting this. A screenshot is the one artefact in a repository
that nothing checks. Code gets tested, numbers get verified, and the picture at
the top of the README quietly goes on showing whatever was true the day someone
took it. This project had exactly that problem: `build_workbook.py` was fixed to
leave the oldest week's trend cell blank — unknown is not the same as
unchanged — and the committed screenshot carried on displaying `0.0%`, which is
precisely the bug the code comment said had been fixed.

Regenerating them is now one command, so there is no excuse for drift.

Needs LibreOffice (for the spreadsheet) and Playwright (for the email HTML):
    pip install playwright && playwright install chromium
"""

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

WB = Path("weekly-report-template.xlsx")
CSV = Path("data/orders_raw.csv")
SHOTS = Path("screenshots")

# How much of each sheet to show. Enough to make the point, not so much that
# the image becomes unreadable in a README at 100% width.
RAW_ROWS = 14


def soffice() -> str:
    for name in ("soffice", "libreoffice"):
        if (found := shutil.which(name)):
            return found
    mac = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    if mac.exists():
        return str(mac)
    sys.exit("ERROR: LibreOffice not found — see recalc.py for install notes.")


def sheet_to_png(src: Path, out: Path, sheet_name: str, width: int = 1180) -> None:
    """
    Render one sheet: LibreOffice to HTML, then a browser screenshot.

    The obvious route is LibreOffice to PDF, and it is wrong here. PDF export
    paginates, so a report wider than the paper loses its right-hand columns —
    on this workbook that silently cropped three of the five KPIs and the whole
    'vs prior week' column. HTML has no page width to overflow, so the image
    shows the sheet as a person scrolling it would see it.

    Re-saving through openpyxl to set print options is not an option either:
    openpyxl drops cached values on save, and the image would come out blank.
    """
    from playwright.sync_api import sync_playwright

    with tempfile.TemporaryDirectory() as tmp:
        subprocess.run(
            [soffice(), "--headless", "--norestore", "--convert-to",
             "html:HTML (StarCalc)", "--outdir", tmp, str(src)],
            capture_output=True, text=True, timeout=300, check=False)
        html = Path(tmp) / (src.stem + ".html")
        if not html.exists():
            sys.exit(f"ERROR: LibreOffice produced no HTML for {src}. "
                     "Is it already running? Headless refuses a second instance.")

        # Calc writes every sheet into one document, each preceded by an anchor
        # named after the tab. Keep the requested one and drop the rest, so the
        # image is of a sheet rather than of the whole workbook.
        css = """
        <style>
          body { font-family: Arial, Helvetica, sans-serif; background:#fff;
                 margin:0; padding:26px 30px; }
          table { border-collapse: collapse; }
          td, th { padding:2px 6px; }
          a[name] + table { margin-bottom: 0; }
        </style>"""
        doc = html.read_text(encoding="utf-8", errors="replace")
        doc = doc.replace("</head>", css + "</head>")
        html.write_text(doc, encoding="utf-8")

        with sync_playwright() as p:
            b = p.chromium.launch()
            page = b.new_page(viewport={"width": width, "height": 900},
                              device_scale_factor=2)
            page.goto(html.resolve().as_uri())
            page.wait_for_timeout(300)
            # Calc's HTML export wraps the sheets in its own furniture: a
            # sheet-navigation link list, horizontal rules, and a heading per
            # tab. None of that is part of the report, so it goes.
            keep = page.evaluate(
                """(name) => {
                    const tables = [...document.querySelectorAll('table')];
                    const anchors = [...document.querySelectorAll('a[name]')]
                        .map(a => a.getAttribute('name'));
                    let idx = anchors.indexOf(name);
                    if (idx < 0) idx = 0;
                    tables.forEach((t, i) => { if (i !== idx) t.remove(); });
                    document.querySelectorAll('a[name], h1, h2, hr')
                        .forEach(e => e.remove());
                    // Paragraphs that are only sheet links are navigation.
                    document.querySelectorAll('p, div').forEach(e => {
                        if (!e.querySelector('table') &&
                            e.querySelectorAll('a[href^="#"]').length) e.remove();
                    });
                    document.querySelectorAll('a[href^="#"]').forEach(a => a.remove());
                    return tables.length;
                }""", sheet_name)
            if not keep:
                sys.exit(f"ERROR: no tables in the HTML export of {src}.")
            page.wait_for_timeout(200)
            out.parent.mkdir(exist_ok=True)
            # Shoot the table itself rather than the viewport, so the image has
            # no dead margin on the right where the sheet stops.
            target = page.query_selector("table") or page
            target.screenshot(path=str(out))
            b.close()
    print(f"wrote {out}")


def raw_extract() -> Path:
    """
    A small workbook holding just the top of the export, so the 'before' image
    shows the mess without being a 1,835-row wall.
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    df = pd.read_csv(CSV, dtype=str).head(RAW_ROWS)
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Orders"
    ws.append(list(df.columns))
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F6FEB")
        c.alignment = Alignment(horizontal="center")
    for row in df.itertuples(index=False):
        ws.append(["" if (v != v or v is None) else v for v in row])
    for col, w in zip("ABCDEFGHIJK",
                      (10, 18, 11, 30, 6, 11, 11, 12, 12, 26, 11)):
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    tmp = Path("data/_raw_preview.xlsx")
    wb.save(tmp)
    return tmp


def email_png() -> None:
    html = SHOTS / "email-preview.html"
    if not html.exists():
        sys.exit("ERROR: run `python3 render_email_preview.py` first.")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  (skipping email-preview.png — playwright not installed)")
        return
    with sync_playwright() as p:
        b = p.chromium.launch()
        page = b.new_page(viewport={"width": 760, "height": 900},
                          device_scale_factor=2)
        page.goto(html.resolve().as_uri())
        page.wait_for_timeout(400)
        page.screenshot(path=str(SHOTS / "email-preview.png"), full_page=True)
        b.close()
    print(f"wrote {SHOTS / 'email-preview.png'}")


def main() -> int:
    if not WB.exists():
        sys.exit("ERROR: weekly-report-template.xlsx not found. "
                 "Run build_workbook.py then recalc.py first.")

    from openpyxl import load_workbook
    if load_workbook(WB, data_only=True)["Weekly Report"]["B6"].value is None:
        sys.exit("ERROR: the workbook has no calculated values, so the report "
                 "would render blank.\n  Run `python3 recalc.py` first.")

    sheet_to_png(WB, SHOTS / "weekly-report.png", "Weekly Report", width=1180)

    preview = raw_extract()
    sheet_to_png(preview, SHOTS / "raw-orders-before.png", "Raw Orders", width=1320)
    preview.unlink(missing_ok=True)

    email_png()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
